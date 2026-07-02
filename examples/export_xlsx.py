#!/usr/bin/env python3
"""Export full agibot_x2 ontology to a multi-sheet Excel."""
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from roboonto.api.loader import OntologyLoader

CATEGORY_COLOR = {
    "hardware":  "FF6B6B",
    "behavior":  "FFD54F",
    "event":     "FF9800",
    "interface": "4DD0E1",
    "action":    "EC407A",
    "meta":      "90A4AE",
}

TYPE_TO_CATEGORY = {
    "ComputeUnit": "hardware", "Joint": "hardware", "Link": "hardware",
    "Sensor": "hardware", "PowerSubsystem": "hardware", "RobotBody": "hardware",
    "EndEffector": "hardware",
    "Topic": "interface", "Service": "interface",
    "MsgSchema": "interface", "SrvSchema": "interface",
    "Mode": "behavior", "PresetMotion": "behavior", "ControlLoop": "behavior",
    "StatusBit": "event", "FaultCode": "event", "TouchEvent": "event",
    "CoordinateFrame": "frame",
    "InputSource": "meta", "PriorityLevel": "meta", "SafetyClass": "meta",
}

# 每类对象的显示列(有序,简明)
COLUMNS_BY_TYPE = {
    "Joint": ["id", "name_human", "urdf_name", "j_label", "dof_axis",
              "position_limit", "velocity_limit", "effort_limit",
              "parent_link", "child_link", "axis_xyz", "source"],
    "Link": ["id", "name_urdf", "mass", "visual_mesh", "source"],
    "Sensor": ["id", "sensor_kind", "vendor", "model", "mount_link", "specs", "source"],
    "ComputeUnit": ["id", "role", "soc", "ai_tops", "memory_gb", "storage_gb",
                    "network_address", "os", "notes", "source"],
    "PowerSubsystem": ["id", "battery_wh", "battery_voltage_nominal",
                       "charging_input_v", "endurance_walking_hours", "source"],
    "RobotBody": ["id", "height_m", "weight_kg", "total_active_dof",
                  "max_speed_m_s", "payload_kg_max_specific_pose", "source"],
    "EndEffector": ["id", "ee_kind", "mount_joint", "dof", "payload_kg", "source"],

    "Topic": ["id", "name", "msg_type", "qos", "frequency_hz",
              "bandwidth_mb_s", "cross_unit_safe", "compute_unit", "notes", "source"],
    "Service": ["id", "name", "srv_type", "cross_unit_reliable", "source"],
    "MsgSchema": ["id", "name", "fields", "fields_grouped", "source"],
    "SrvSchema": ["id", "name", "source"],

    "Mode": ["id", "mode_value", "name_human", "description", "use_case", "safety_note", "source"],
    "PresetMotion": ["id", "name_human", "motion_value", "area_value",
                     "required_mode", "notes", "source"],
    "ControlLoop": ["id", "name_human", "description", "frequency_hz",
                    "arbitrates_topics", "source"],

    "StatusBit": ["id", "carrier_field", "bit_index", "bit_name", "severity",
                  "semantic", "source"],
    "FaultCode": ["id", "carrier", "code", "severity", "semantic", "source"],
    "TouchEvent": ["id", "event_code", "semantic", "source"],

    "InputSource": ["id", "source_name", "display_name", "priority",
                    "timeout_ms", "requires_registration", "origin", "source"],
    "PriorityLevel": ["id", "value", "name", "description", "use_case", "preemptable", "source"],

    "Action": ["id", "description", "invoker_type", "invoker_name",
               "parameters", "preconditions", "affects", "safety_class",
               "required_input_source", "idempotent", "source"],
}


def format_value(v, max_len=120):
    if v is None:
        return ""
    if isinstance(v, dict):
        s = ", ".join(f"{k}={v[k]}" for k in v)
    elif isinstance(v, list):
        s = "; ".join(format_value(x, max_len=40) for x in v)
    else:
        s = str(v)
    if len(s) > max_len:
        s = s[:max_len - 1] + "…"
    return s


def extract_row(obj, columns, is_action=False):
    """Convert an object or action dict to a row dict keyed by column name."""
    row = {}
    if is_action:
        row["id"] = obj.get("type_id")
        row["description"] = obj.get("description", "")
        inv = obj.get("invoker") or {}
        row["invoker_type"] = inv.get("type", "")
        row["invoker_name"] = inv.get("name", "") or inv.get("name_template", "")
        params = obj.get("parameters") or []
        row["parameters"] = "; ".join(
            f"{p.get('name')}:{p.get('type','?')}"
            + ("(req)" if p.get("required") else "")
            for p in params
        )
        pres = obj.get("preconditions") or []
        row["preconditions"] = f"{len(pres)} rules" if pres else ""
        affects = obj.get("affects") or []
        row["affects"] = "; ".join(affects)
        row["safety_class"] = obj.get("safety_class", "")
        row["required_input_source"] = obj.get("required_input_source", "")
        row["idempotent"] = obj.get("idempotent", "")
        src = obj.get("source") or {}
        row["source"] = src.get("locator", "")
        return row

    row["id"] = obj.get("id")
    props = obj.get("properties") or {}
    src = obj.get("source") or {}
    for col in columns:
        if col == "id":
            continue
        if col == "source":
            row[col] = src.get("locator", "")
            continue
        row[col] = format_value(props.get(col))
    return row


def write_sheet(wb, title, objs, columns, category_hex, is_action=False):
    ws = wb.create_sheet(title)

    # Header
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", start_color=category_hex)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    center = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for ci, col in enumerate(columns, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border

    # Freeze header row
    ws.freeze_panes = "A2"

    # Rows
    body_font = Font(name="Arial", size=10)
    alt_fill = PatternFill("solid", start_color="F8F9FA")

    for ri, obj in enumerate(objs, 2):
        row = extract_row(obj, columns, is_action=is_action)
        for ci, col in enumerate(columns, 1):
            val = row.get(col, "")
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = body_font
            c.border = thin_border
            c.alignment = Alignment(
                horizontal="left", vertical="top",
                wrap_text=True,
            )
            if ri % 2 == 0:
                c.fill = alt_fill

    # Column widths
    widths = {
        "id": 42, "source": 40, "name_human": 22, "description": 50,
        "name": 45, "specs": 50, "fields": 60, "fields_grouped": 60,
        "notes": 40, "semantic": 30, "visual_mesh": 30, "soc": 35,
        "parameters": 40, "preconditions": 12, "affects": 40,
        "invoker_name": 40, "invoker_type": 14,
        "position_limit": 30, "velocity_limit": 22, "effort_limit": 20,
        "parent_link": 42, "child_link": 42, "mount_link": 30,
        "compute_unit": 28, "arbitrates_topics": 35, "required_mode": 36,
        "description": 40, "use_case": 25, "safety_note": 40,
        "carrier_field": 18, "bit_name": 32, "severity": 10,
        "ee_kind": 12, "mount_joint": 30, "msg_type": 40, "srv_type": 40,
        "origin": 14, "display_name": 22, "source_name": 14,
        "axis_xyz": 14, "urdf_name": 28, "dof_axis": 10, "j_label": 8,
        "bit_index": 9, "severity": 10, "code": 8, "event_code": 10,
        "value": 7, "priority": 10, "timeout_ms": 12,
        "cross_unit_safe": 15, "cross_unit_reliable": 18,
        "qos": 16, "frequency_hz": 14, "bandwidth_mb_s": 14,
        "motion_value": 13, "area_value": 10, "preemptable": 12,
        "requires_registration": 20,
        "ai_tops": 9, "memory_gb": 10, "storage_gb": 10,
        "mass": 8, "role": 13, "os": 16, "network_address": 15,
        "battery_wh": 11, "battery_voltage_nominal": 12,
        "charging_input_v": 18, "endurance_walking_hours": 10,
        "height_m": 9, "weight_kg": 9, "total_active_dof": 10,
        "max_speed_m_s": 10, "payload_kg_max_specific_pose": 14,
        "dof": 6, "payload_kg": 10, "carrier": 22,
        "safety_class": 12, "required_input_source": 14, "idempotent": 10,
    }
    for ci, col in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 18)

    ws.row_dimensions[1].height = 22


def main():
    loader = OntologyLoader(Path("robots/agibot_x2"))
    ont = loader.load()

    wb = Workbook()
    wb.remove(wb.active)  # remove default

    # -------- Overview sheet first --------
    ws = wb.create_sheet("Overview", 0)

    ws["A1"] = "RoboOnto · AgiBot X2 Ontology"
    ws["A1"].font = Font(name="Arial", bold=True, size=18, color="0F1420")
    ws.merge_cells("A1:E1")

    ws["A2"] = f"v0.2 · {ont.robot_meta.get('model', '')}"
    ws["A2"].font = Font(name="Arial", italic=True, size=11, color="6B7590")
    ws.merge_cells("A2:E2")

    # Stats table
    ws["A4"] = "Summary"
    ws["A4"].font = Font(name="Arial", bold=True, size=13)

    headers = ["Category", "Object Types", "Instance Count", "Sheet"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=6, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="0F1420")
        c.alignment = Alignment(horizontal="left")

    # Group by category
    cat_data = {}
    for t, items in ont.objects_by_type.items():
        cat = TYPE_TO_CATEGORY.get(t, "meta")
        cat_data.setdefault(cat, {"types": [], "count": 0})
        cat_data[cat]["types"].append(f"{t}({len(items)})")
        cat_data[cat]["count"] += len(items)
    cat_data["action"] = {"types": [f"Action({len(ont.actions_by_id)})"], "count": len(ont.actions_by_id)}

    row_idx = 7
    category_order = ["hardware", "behavior", "event", "interface", "action", "meta"]
    for cat in category_order:
        if cat not in cat_data:
            continue
        data = cat_data[cat]
        color = CATEGORY_COLOR[cat]
        ws.cell(row=row_idx, column=1, value=cat)
        ws.cell(row=row_idx, column=1).fill = PatternFill("solid", start_color=color)
        ws.cell(row=row_idx, column=1).font = Font(name="Arial", bold=True, color="FFFFFF")
        ws.cell(row=row_idx, column=2, value=", ".join(data["types"]))
        ws.cell(row=row_idx, column=3, value=data["count"])
        ws.cell(row=row_idx, column=4, value=f"→ see sheets with prefix '{cat}.'")
        for col in range(1, 5):
            ws.cell(row=row_idx, column=col).font = Font(name="Arial", size=10)
            ws.cell(row=row_idx, column=col).alignment = Alignment(vertical="center", wrap_text=True)
        row_idx += 1

    # Totals row
    ws.cell(row=row_idx, column=1, value="TOTAL")
    ws.cell(row=row_idx, column=1).font = Font(name="Arial", bold=True)
    ws.cell(row=row_idx, column=3, value=f"=SUM(C7:C{row_idx-1})")
    ws.cell(row=row_idx, column=3).font = Font(name="Arial", bold=True)

    # Column widths for overview
    widths = [18, 60, 16, 40]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28
    for r in range(6, row_idx + 1):
        ws.row_dimensions[r].height = 20

    # Legend note
    note_row = row_idx + 2
    ws.cell(row=note_row, column=1, value="Sheet naming: <category>.<Type>  —  e.g., hardware.Joint, event.StatusBit, action.Action").font = Font(name="Arial", italic=True, color="6B7590")
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)

    ws.cell(row=note_row + 1, column=1, value="Column meaning: 'source' = 文档或 URDF 来源；'position_limit' = URDF 权威范围；'effort_limit' = 各关节真实扭矩档位").font = Font(name="Arial", italic=True, color="6B7590")
    ws.merge_cells(start_row=note_row + 1, start_column=1, end_row=note_row + 1, end_column=5)

    # -------- Per-type sheets, grouped by category --------
    for cat in category_order:
        if cat == "action":
            # Special: just one sheet for all actions
            actions = sorted(ont.actions_by_id.values(), key=lambda a: a.get("type_id", ""))
            write_sheet(wb, "action.Action", actions,
                        COLUMNS_BY_TYPE["Action"], CATEGORY_COLOR["action"],
                        is_action=True)
            continue

        # For other categories: one sheet per object type
        types_in_cat = [t for t, items in ont.objects_by_type.items()
                        if TYPE_TO_CATEGORY.get(t) == cat]
        for t in sorted(types_in_cat):
            items = ont.objects_by_type[t]
            if not items:
                continue
            cols = COLUMNS_BY_TYPE.get(t, ["id", "source"])
            sheet_name = f"{cat}.{t}"
            # Excel sheet name max 31 chars
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
            # Sort by id for stability
            items = sorted(items, key=lambda o: o.get("id", ""))
            write_sheet(wb, sheet_name, items, cols, CATEGORY_COLOR[cat])

    # Save
    out = Path(__file__).parent / "ontology_inventory.xlsx"
    wb.save(out)
    print(f"Wrote {out}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
